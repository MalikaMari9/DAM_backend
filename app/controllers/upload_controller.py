from sqlalchemy.orm import Session
from fastapi import HTTPException, status
from bson import ObjectId
from datetime import datetime
from pymongo.errors import DuplicateKeyError
import csv
import io
import json
from pathlib import Path
from app.models.account_model import Account
from app.models.enums import AccountRole, DataDomain, UploadStatus
from app.core.mongo import get_imhe_collection, get_openaq_collection
from app.core.db import SessionLocal
from app.core.config import get_settings
from app.repositories.org_repo import get_org_by_id
from app.repositories.upload_repo import (
    create_upload,
    list_uploads,
    list_uploads_by_org,
    get_upload_by_id,
    update_upload_status,
    delete_upload,
)
from app.schemas.upload_schema import (
    UploadCreate,
    UploadUpdateStatus,
    HealthIMHERecordManual,
    PollutionOpenAQRecordManual,
    UploadRecordUpdate,
)

IMHE_REQUIRED_FIELDS = [
    "population_group_id",
    "population_group_name",
    "measure_id",
    "measure_name",
    "location_id",
    "location_name",
    "sex_id",
    "sex_name",
    "age_id",
    "age_name",
    "cause_id",
    "cause_name",
    "metric_id",
    "metric_name",
    "year",
    "val",
    "upper",
    "lower",
]

_CSV_VALIDATION_CACHE: dict[str, dict] = {}
_CSV_CACHE_MAX_AGE_SECONDS = 60 * 30

POLLUTION_REQUIRED_FIELDS = [
    "location_name",
    "pollutant",
    "units",
    "year",
    "value",
]

_SUPPORTED_UPLOAD_EXTS = {".csv", ".xlsx", ".xls", ".json"}


def _get_file_ext(filename: str | None) -> str:
    if not filename:
        return ""
    return Path(filename).suffix.lower()


def _normalize_units(units: str | None) -> str | None:
    if not isinstance(units, str):
        return units
    return (
        units.replace("Âµg/mÂ³", "µg/m³")
        .replace("Âµg/m3", "µg/m³")
        .replace("Ã‚Âµg/mÃ‚Â³", "µg/m³")
        .replace("Ã‚Âµg/m3", "µg/m³")
    )


def _rows_from_json(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig")
    data = json.loads(text)
    if isinstance(data, dict):
        for key in ("items", "records", "data"):
            if isinstance(data.get(key), list):
                data = data[key]
                break
    if not isinstance(data, list):
        raise ValueError("JSON must be an array of objects.")
    rows = [row for row in data if isinstance(row, dict)]
    if not rows:
        raise ValueError("JSON contains no data rows.")
    return rows


def _rows_from_excel(file_bytes: bytes, ext: str) -> list[dict]:
    if ext == ".xlsx":
        try:
            import openpyxl
        except Exception as exc:
            raise ValueError("Excel .xlsx support requires openpyxl.") from exc
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("Excel file is missing headers.")
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        if not any(headers):
            raise ValueError("Excel file has empty headers.")
        rows: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        if not rows:
            raise ValueError("Excel file contains no data rows.")
        return rows

    if ext == ".xls":
        try:
            import xlrd
        except Exception as exc:
            raise ValueError("Excel .xls support requires xlrd.") from exc
        book = xlrd.open_workbook(file_contents=file_bytes)
        sheet = book.sheet_by_index(0)
        if sheet.nrows < 1:
            raise ValueError("Excel file is missing headers.")
        headers = [str(h).strip() if h is not None else "" for h in sheet.row_values(0)]
        if not any(headers):
            raise ValueError("Excel file has empty headers.")
        rows: list[dict] = []
        for r in range(1, sheet.nrows):
            values = sheet.row_values(r)
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
        if not rows:
            raise ValueError("Excel file contains no data rows.")
        return rows

    raise ValueError("Unsupported Excel format.")


def _cache_cleanup():
    now = datetime.utcnow().timestamp()
    expired = [k for k, v in _CSV_VALIDATION_CACHE.items() if now - v["created_at"] > _CSV_CACHE_MAX_AGE_SECONDS]
    for k in expired:
        _CSV_VALIDATION_CACHE.pop(k, None)


def _key_tuple(doc: dict) -> tuple:
    return (
        doc.get("population_group_id"),
        doc.get("measure_id"),
        doc.get("location_id"),
        doc.get("sex_id"),
        doc.get("age_id"),
        doc.get("cause_id"),
        doc.get("metric_id"),
        doc.get("year"),
    )


def _find_existing_keys(col, keys: list[tuple]) -> set[tuple]:
    existing: set[tuple] = set()
    chunk = 500
    for i in range(0, len(keys), chunk):
        batch = keys[i:i + chunk]
        or_filters = [
            {
                "population_group_id": k[0],
                "measure_id": k[1],
                "location_id": k[2],
                "sex_id": k[3],
                "age_id": k[4],
                "cause_id": k[5],
                "metric_id": k[6],
                "year": k[7],
            }
            for k in batch
        ]
        cursor = col.find(
            {"$or": or_filters},
            {
                "population_group_id": 1,
                "measure_id": 1,
                "location_id": 1,
                "sex_id": 1,
                "age_id": 1,
                "cause_id": 1,
                "metric_id": 1,
                "year": 1,
            },
        )
        for doc in cursor:
            existing.add(
                (
                    doc.get("population_group_id"),
                    doc.get("measure_id"),
                    doc.get("location_id"),
                    doc.get("sex_id"),
                    doc.get("age_id"),
                    doc.get("cause_id"),
                    doc.get("metric_id"),
                    doc.get("year"),
                )
            )
    return existing


def _normalize_country_name(value: str) -> str:
    return " ".join(value.strip().lower().split())


def _parse_optional_float(value: str | None):
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    return float(text)


def _parse_optional_int(value: str | None):
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    return int(text)


def _require_fields(rows: list[dict], required: list[str]):
    if not rows:
        raise ValueError("File contains no data rows.")
    headers = set(rows[0].keys())
    missing = [field for field in required if field not in headers]
    if missing:
        raise ValueError(f"File is missing required columns: {', '.join(missing)}")


def _parse_imhe_rows(rows: list[dict], expected_country: str, row_offset: int) -> tuple[list[dict], str]:
    _require_fields(rows, IMHE_REQUIRED_FIELDS)
    docs: list[dict] = []
    expected_norm = _normalize_country_name(expected_country)
    location_label: str | None = None

    for idx, row in enumerate(rows, start=row_offset):
        location_name = (row.get("location_name") or "").strip()
        if not location_name:
            raise ValueError(f"Row {idx}: location_name is required.")
        if _normalize_country_name(location_name) != expected_norm:
            raise ValueError(
                f"Row {idx}: location_name '{location_name}' does not match org country '{expected_country}'."
            )
        if location_label is None:
            location_label = location_name

        try:
            doc = {
                "population_group_id": int(str(row["population_group_id"]).strip()),
                "population_group_name": (row.get("population_group_name") or "").strip(),
                "measure_id": int(str(row["measure_id"]).strip()),
                "measure_name": (row.get("measure_name") or "").strip(),
                "location_id": int(str(row["location_id"]).strip()),
                "location_name": location_name,
                "sex_id": int(str(row["sex_id"]).strip()),
                "sex_name": (row.get("sex_name") or "").strip(),
                "age_id": int(str(row["age_id"]).strip()),
                "age_name": (row.get("age_name") or "").strip(),
                "cause_id": int(str(row["cause_id"]).strip()),
                "cause_name": (row.get("cause_name") or "").strip(),
                "metric_id": int(str(row["metric_id"]).strip()),
                "metric_name": (row.get("metric_name") or "").strip(),
                "year": int(str(row["year"]).strip()),
                "val": float(str(row["val"]).strip()),
                "upper": float(str(row["upper"]).strip()),
                "lower": float(str(row["lower"]).strip()),
            }
        except (ValueError, TypeError) as exc:
            raise ValueError(f"Row {idx}: invalid numeric value ({exc}).") from exc

        docs.append(doc)

    if not docs:
        raise ValueError("File contains no data rows.")

    return docs, location_label or expected_country


def _parse_pollution_rows(rows: list[dict], expected_country: str, row_offset: int) -> tuple[list[dict], str]:
    _require_fields(rows, POLLUTION_REQUIRED_FIELDS)
    docs: list[dict] = []
    expected_norm = _normalize_country_name(expected_country)
    location_label: str | None = None

    for idx, row in enumerate(rows, start=row_offset):
        location_name = (row.get("location_name") or "").strip()
        if not location_name:
            raise ValueError(f"Row {idx}: location_name is required.")
        if location_label is None:
            location_label = location_name

        country_name = (row.get("country_name") or expected_country).strip()
        if _normalize_country_name(country_name) != expected_norm:
            raise ValueError(
                f"Row {idx}: country_name '{country_name}' does not match org country '{expected_country}'."
            )

        pollutant = (row.get("pollutant") or "").strip()
        if not pollutant:
            raise ValueError(f"Row {idx}: pollutant is required.")

        units = _normalize_units((row.get("units") or "µg/m³").strip())
        if not units:
            raise ValueError(f"Row {idx}: units is required.")

        try:
            doc = {
                "country_name": country_name,
                "year": int(str(row["year"]).strip()),
                "location_name": location_name,
                "pollutant": pollutant,
                "units": units,
                "value": float(str(row["value"]).strip()),
                "latitude": _parse_optional_float(row.get("latitude")),
                "longitude": _parse_optional_float(row.get("longitude")),
                "min": _parse_optional_float(row.get("min")),
                "max": _parse_optional_float(row.get("max")),
                "median": _parse_optional_float(row.get("median")),
                "avg": _parse_optional_float(row.get("avg")),
                "coverage_percent": _parse_optional_float(row.get("coverage_percent")),
                "sensor_id": _parse_optional_int(row.get("sensor_id")),
                "location_id": _parse_optional_int(row.get("location_id")),
            }
        except (ValueError, TypeError) as exc:
            raise ValueError(f"Row {idx}: invalid numeric value ({exc}).") from exc

        docs.append(doc)

    if not docs:
        raise ValueError("File contains no data rows.")

    return docs, location_label or expected_country


def _pollution_key_tuple(doc: dict) -> tuple:
    return (
        doc.get("country_name"),
        doc.get("location_name"),
        doc.get("pollutant"),
        doc.get("year"),
    )


def _find_existing_pollution_keys(col, keys: list[tuple]) -> set[tuple]:
    existing: set[tuple] = set()
    chunk = 500
    for i in range(0, len(keys), chunk):
        batch = keys[i:i + chunk]
        or_filters = [
            {
                "country_name": k[0],
                "location_name": k[1],
                "pollutant": k[2],
                "year": k[3],
            }
            for k in batch
        ]
        cursor = col.find(
            {"$or": or_filters},
            {
                "country_name": 1,
                "location_name": 1,
                "pollutant": 1,
                "year": 1,
            },
        )
        for doc in cursor:
            existing.add(
                (
                    doc.get("country_name"),
                    doc.get("location_name"),
                    doc.get("pollutant"),
                    doc.get("year"),
                )
            )
    return existing


def _parse_imhe_csv(file_bytes: bytes, expected_country: str) -> tuple[list[dict], str]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV file is missing headers.")
    rows = list(reader)
    if not rows:
        raise ValueError("CSV contains no data rows.")
    return _parse_imhe_rows(rows, expected_country, row_offset=2)


def _parse_pollution_csv(file_bytes: bytes, expected_country: str) -> tuple[list[dict], str]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV file is missing headers.")

    missing = [field for field in POLLUTION_REQUIRED_FIELDS if field not in reader.fieldnames]
    if missing:
        raise ValueError(f"CSV is missing required columns: {', '.join(missing)}")

    docs: list[dict] = []
    expected_norm = _normalize_country_name(expected_country)
    location_label: str | None = None

    for row_index, row in enumerate(reader, start=2):
        location_name = (row.get("location_name") or "").strip()
        if not location_name:
            raise ValueError(f"Row {row_index}: location_name is required.")
        if location_label is None:
            location_label = location_name

        country_name = (row.get("country_name") or expected_country).strip()
        if _normalize_country_name(country_name) != expected_norm:
            raise ValueError(
                f"Row {row_index}: country_name '{country_name}' does not match org country '{expected_country}'."
            )

        pollutant = (row.get("pollutant") or "").strip()
        if not pollutant:
            raise ValueError(f"Row {row_index}: pollutant is required.")

        units = (row.get("units") or "µg/m³").strip()
        if not units:
            raise ValueError(f"Row {row_index}: units is required.")

        try:
            doc = {
                "country_name": country_name,
                "year": int(row["year"]),
                "location_name": location_name,
                "pollutant": pollutant,
                "units": units,
                "value": float(row["value"]),
                "latitude": _parse_optional_float(row.get("latitude")),
                "longitude": _parse_optional_float(row.get("longitude")),
                "min": _parse_optional_float(row.get("min")),
                "max": _parse_optional_float(row.get("max")),
                "median": _parse_optional_float(row.get("median")),
                "avg": _parse_optional_float(row.get("avg")),
                "coverage_percent": _parse_optional_float(row.get("coverage_percent")),
                "sensor_id": _parse_optional_int(row.get("sensor_id")),
                "location_id": _parse_optional_int(row.get("location_id")),
            }
        except (ValueError, TypeError) as exc:
            raise ValueError(f"Row {row_index}: invalid numeric value ({exc}).") from exc

        docs.append(doc)

    if not docs:
        raise ValueError("CSV contains no data rows.")

    return docs, location_label or expected_country


def _parse_pollution_csv(file_bytes: bytes, expected_country: str) -> tuple[list[dict], str]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV file is missing headers.")
    rows = list(reader)
    if not rows:
        raise ValueError("CSV contains no data rows.")
    return _parse_pollution_rows(rows, expected_country, row_offset=2)


def _parse_imhe_json(file_bytes: bytes, expected_country: str) -> tuple[list[dict], str]:
    rows = _rows_from_json(file_bytes)
    return _parse_imhe_rows(rows, expected_country, row_offset=1)


def _parse_pollution_json(file_bytes: bytes, expected_country: str) -> tuple[list[dict], str]:
    rows = _rows_from_json(file_bytes)
    return _parse_pollution_rows(rows, expected_country, row_offset=1)


def _parse_imhe_excel(file_bytes: bytes, expected_country: str, ext: str) -> tuple[list[dict], str]:
    rows = _rows_from_excel(file_bytes, ext)
    return _parse_imhe_rows(rows, expected_country, row_offset=2)


def _parse_pollution_excel(file_bytes: bytes, expected_country: str, ext: str) -> tuple[list[dict], str]:
    rows = _rows_from_excel(file_bytes, ext)
    return _parse_pollution_rows(rows, expected_country, row_offset=2)


def _parse_imhe_upload(file_bytes: bytes, filename: str | None, expected_country: str) -> tuple[list[dict], str]:
    ext = _get_file_ext(filename)
    if ext == ".csv":
        return _parse_imhe_csv(file_bytes, expected_country)
    if ext == ".json":
        return _parse_imhe_json(file_bytes, expected_country)
    if ext in (".xlsx", ".xls"):
        return _parse_imhe_excel(file_bytes, expected_country, ext)
    raise ValueError("Unsupported file type. Please upload CSV, Excel (.xlsx/.xls), or JSON.")


def _parse_pollution_upload(file_bytes: bytes, filename: str | None, expected_country: str) -> tuple[list[dict], str]:
    ext = _get_file_ext(filename)
    if ext == ".csv":
        return _parse_pollution_csv(file_bytes, expected_country)
    if ext == ".json":
        return _parse_pollution_json(file_bytes, expected_country)
    if ext in (".xlsx", ".xls"):
        return _parse_pollution_excel(file_bytes, expected_country, ext)
    raise ValueError("Unsupported file type. Please upload CSV, Excel (.xlsx/.xls), or JSON.")


def _process_imhe_csv_upload(upload_id: int, batch_id: str, filename: str, file_bytes: bytes, org_country: str):
    if SessionLocal is None:
        raise RuntimeError("DATABASE_URL is not set")
    db = SessionLocal()
    try:
        docs, _location = _parse_imhe_upload(file_bytes, filename, org_country)
        col = get_imhe_collection()
        batch_obj = ObjectId(batch_id)
        for doc in docs:
            doc["_source_batch"] = batch_obj
            doc["_source_file"] = filename
        col.insert_many(docs, ordered=False)
        upload = get_upload_by_id(db, upload_id)
        if upload:
            update_upload_status(
                db,
                upload,
                UploadUpdateStatus(status=UploadStatus.PROCESSED),
            )
    except Exception as exc:
        message = str(exc)
        upload = get_upload_by_id(db, upload_id)
        if upload:
            update_upload_status(
                db,
                upload,
                UploadUpdateStatus(status=UploadStatus.FAILED, error_message=message[:1000]),
            )
    finally:
        db.close()


def create_org_upload(db: Session, account: Account, data: UploadCreate):
    if account.role != AccountRole.ORG or not account.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Organization access required")
    org = get_org_by_id(db, account.org_id)
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Organization not found")

    return create_upload(
        db,
        account_id=account.account_id,
        org_id=org.org_id,
        data_domain=org.data_domain,
        country=org.country,
        data=data,
    )


def create_health_csv_validation(
    db: Session,
    account: Account,
    file_bytes: bytes,
    filename: str,
):
    if account.role != AccountRole.ORG or not account.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Organization access required")
    org = get_org_by_id(db, account.org_id)
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Organization not found")
    if org.data_domain != DataDomain.HEALTH:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Health organization access required")

    settings = get_settings()
    if len(file_bytes) > settings.max_upload_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Max size is {settings.max_upload_bytes} bytes.",
        )

    docs, _location = _parse_imhe_upload(file_bytes, filename, org.country)
    col = get_imhe_collection()
    keys = [_key_tuple(doc) for doc in docs]

    existing_keys = _find_existing_keys(col, list(set(keys)))
    seen: set[tuple] = set()
    dupes: list[tuple] = []
    new_docs: list[dict] = []

    for doc in docs:
        key = _key_tuple(doc)
        if key in existing_keys or key in seen:
            dupes.append(key)
        else:
            seen.add(key)
            new_docs.append(doc)

    token = str(ObjectId())
    _cache_cleanup()
    _CSV_VALIDATION_CACHE[token] = {
        "created_at": datetime.utcnow().timestamp(),
        "docs": new_docs,
        "filename": filename,
        "org_id": org.org_id,
        "country": org.country,
        "dupes": dupes,
        "domain": "health",
    }
    dupe_samples = [
        {
            "population_group_id": k[0],
            "measure_id": k[1],
            "location_id": k[2],
            "sex_id": k[3],
            "age_id": k[4],
            "cause_id": k[5],
            "metric_id": k[6],
            "year": k[7],
        }
        for k in dupes[:5]
    ]
    return {
        "token": token,
        "total_rows": len(docs),
        "dupe_rows": len(dupes),
        "new_rows": len(new_docs),
        "dupe_samples": dupe_samples,
        "dupe_total": len(dupes),
        "token_expires_seconds": _CSV_CACHE_MAX_AGE_SECONDS,
    }


def create_pollution_csv_validation(
    db: Session,
    account: Account,
    file_bytes: bytes,
    filename: str,
):
    if account.role != AccountRole.ORG or not account.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Organization access required")
    org = get_org_by_id(db, account.org_id)
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Organization not found")
    if org.data_domain != DataDomain.POLLUTION:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Pollution organization access required")

    settings = get_settings()
    if len(file_bytes) > settings.max_upload_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Max size is {settings.max_upload_bytes} bytes.",
        )

    docs, _location = _parse_pollution_upload(file_bytes, filename, org.country)
    col = get_openaq_collection()
    keys = [_pollution_key_tuple(doc) for doc in docs]

    existing_keys = _find_existing_pollution_keys(col, list(set(keys)))
    seen: set[tuple] = set()
    dupes: list[tuple] = []
    new_docs: list[dict] = []

    for doc in docs:
        key = _pollution_key_tuple(doc)
        if key in existing_keys or key in seen:
            dupes.append(key)
        else:
            seen.add(key)
            new_docs.append(doc)

    token = str(ObjectId())
    _cache_cleanup()
    _CSV_VALIDATION_CACHE[token] = {
        "created_at": datetime.utcnow().timestamp(),
        "docs": new_docs,
        "filename": filename,
        "org_id": org.org_id,
        "country": org.country,
        "dupes": dupes,
        "domain": "pollution",
    }
    dupe_samples = [
        {
            "country_name": k[0],
            "location_name": k[1],
            "pollutant": k[2],
            "year": k[3],
        }
        for k in dupes[:5]
    ]
    return {
        "token": token,
        "total_rows": len(docs),
        "dupe_rows": len(dupes),
        "new_rows": len(new_docs),
        "dupe_samples": dupe_samples,
        "dupe_total": len(dupes),
        "token_expires_seconds": _CSV_CACHE_MAX_AGE_SECONDS,
    }


def list_csv_dupes(db: Session, account: Account, token: str, limit: int, offset: int):
    _cache_cleanup()
    payload = _CSV_VALIDATION_CACHE.get(token)
    if not payload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload token expired")
    if account.role != AccountRole.ORG or not account.org_id or account.org_id != payload["org_id"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")

    dupes = payload.get("dupes", [])
    total = len(dupes)
    slice_dupes = dupes[int(offset): int(offset) + int(limit)]
    items = [
        {
            "population_group_id": k[0],
            "measure_id": k[1],
            "location_id": k[2],
            "sex_id": k[3],
            "age_id": k[4],
            "cause_id": k[5],
            "metric_id": k[6],
            "year": k[7],
        }
        for k in slice_dupes
    ]
    return {"total": total, "items": items}


def list_pollution_csv_dupes(db: Session, account: Account, token: str, limit: int, offset: int):
    _cache_cleanup()
    payload = _CSV_VALIDATION_CACHE.get(token)
    if not payload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload token expired")
    if account.role != AccountRole.ORG or not account.org_id or account.org_id != payload["org_id"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")
    if payload.get("domain") != "pollution":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Token is not for pollution uploads")

    dupes = payload.get("dupes", [])
    total = len(dupes)
    slice_dupes = dupes[int(offset): int(offset) + int(limit)]
    items = [
        {
            "country_name": k[0],
            "location_name": k[1],
            "pollutant": k[2],
            "year": k[3],
        }
        for k in slice_dupes
    ]
    return {"total": total, "items": items}


def confirm_health_csv_upload(db: Session, account: Account, token: str):
    _cache_cleanup()
    payload = _CSV_VALIDATION_CACHE.pop(token, None)
    if not payload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload token expired")
    if account.role != AccountRole.ORG or not account.org_id or account.org_id != payload["org_id"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")

    col = get_imhe_collection()
    batch_id = str(ObjectId())
    batch_obj = ObjectId(batch_id)
    docs = payload["docs"]
    for doc in docs:
        doc["_source_batch"] = batch_obj
        doc["_source_file"] = payload["filename"]

    if docs:
        try:
            col.insert_many(docs, ordered=False)
        except DuplicateKeyError:
            # Ignore dupes inserted between validate/confirm
            pass

    upload = create_upload(
        db,
        account_id=account.account_id,
        org_id=payload["org_id"],
        data_domain=DataDomain.HEALTH,
        country=payload["country"],
        data=UploadCreate(mongo_collection="IMHE", mongo_ref_id=batch_id),
    )
    update_upload_status(db, upload, UploadUpdateStatus(status=UploadStatus.PROCESSED))
    return upload


def confirm_pollution_csv_upload(db: Session, account: Account, token: str):
    _cache_cleanup()
    payload = _CSV_VALIDATION_CACHE.pop(token, None)
    if not payload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload token expired")
    if account.role != AccountRole.ORG or not account.org_id or account.org_id != payload["org_id"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")
    if payload.get("domain") != "pollution":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Token is not for pollution uploads")

    col = get_openaq_collection()
    batch_id = str(ObjectId())
    batch_obj = ObjectId(batch_id)
    docs = payload["docs"]
    for doc in docs:
        doc["_source_batch"] = batch_obj
        doc["_source_file"] = payload["filename"]

    if docs:
        try:
            col.insert_many(docs, ordered=False)
        except DuplicateKeyError:
            pass

    upload = create_upload(
        db,
        account_id=account.account_id,
        org_id=payload["org_id"],
        data_domain=DataDomain.POLLUTION,
        country=payload["country"],
        data=UploadCreate(mongo_collection="OpenAQ", mongo_ref_id=batch_id),
    )
    update_upload_status(db, upload, UploadUpdateStatus(status=UploadStatus.PROCESSED))
    return upload


def _resolve_id(col, id_field: str, name_field: str, name_value: str) -> int:
    doc = col.aggregate(
        [
            {"$match": {name_field: name_value}},
            {"$group": {"_id": f"${id_field}", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 1},
        ]
    )
    result = list(doc)
    if result:
        return int(result[0]["_id"])

    max_doc = list(
        col.aggregate(
            [
                {"$match": {id_field: {"$type": "number"}}},
                {"$group": {"_id": None, "max_id": {"$max": f"${id_field}"}}},
            ]
        )
    )
    max_id = int(max_doc[0]["max_id"]) if max_doc and max_doc[0].get("max_id") is not None else 0
    return max_id + 1


def create_health_record_upload(db: Session, account: Account, record: HealthIMHERecordManual):
    if account.role != AccountRole.ORG or not account.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Organization access required")
    org = get_org_by_id(db, account.org_id)
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Organization not found")
    if org.data_domain != DataDomain.HEALTH:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Health organization access required")

    if _normalize_country_name(record.location_name) != _normalize_country_name(org.country):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="location_name does not match organization country",
        )
    # Basic validation
    current_year = datetime.utcnow().year
    if record.year < 1900 or record.year > current_year + 3:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"year must be between 1900 and {current_year + 3}",
        )
    if record.val is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="val is required")
    if record.upper is not None and record.lower is not None:
        if record.lower > record.upper:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="lower cannot exceed upper")

    batch_id = str(ObjectId())
    col = get_imhe_collection()
    doc = record.model_dump()
    doc["population_group_id"] = 1
    doc["population_group_name"] = "All Population"
    doc["measure_id"] = _resolve_id(col, "measure_id", "measure_name", record.measure_name)
    doc["location_id"] = _resolve_id(col, "location_id", "location_name", record.location_name)
    doc["sex_id"] = _resolve_id(col, "sex_id", "sex_name", record.sex_name)
    doc["age_id"] = _resolve_id(col, "age_id", "age_name", record.age_name)
    doc["cause_id"] = _resolve_id(col, "cause_id", "cause_name", record.cause_name)
    doc["metric_id"] = _resolve_id(col, "metric_id", "metric_name", record.metric_name)
    if doc.get("upper") is None:
        doc["upper"] = doc["val"]
    if doc.get("lower") is None:
        doc["lower"] = doc["val"]
    doc["_source_batch"] = ObjectId(batch_id)
    doc["_source_file"] = "manual"
    try:
        col.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Record already exists for the same keys (population/measure/location/sex/age/cause/metric/year).",
        )

    upload = create_upload(
        db,
        account_id=account.account_id,
        org_id=org.org_id,
        data_domain=org.data_domain,
        country=org.country,
        data=UploadCreate(mongo_collection="IMHE", mongo_ref_id=batch_id),
    )
    update_upload_status(db, upload, UploadUpdateStatus(status=UploadStatus.PROCESSED))
    return upload


def create_pollution_record_upload(db: Session, account: Account, record: PollutionOpenAQRecordManual):
    if account.role != AccountRole.ORG or not account.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Organization access required")
    org = get_org_by_id(db, account.org_id)
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Organization not found")
    if org.data_domain != DataDomain.POLLUTION:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Pollution organization access required")

    current_year = datetime.utcnow().year
    if record.year < 1900 or record.year > current_year + 3:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"year must be between 1900 and {current_year + 3}",
        )
    if record.value is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="value is required")

    batch_id = str(ObjectId())
    col = get_openaq_collection()
    doc = record.model_dump()
    doc["country_name"] = org.country
    doc["_source_batch"] = ObjectId(batch_id)
    doc["_source_file"] = "manual"

    try:
        col.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Record already exists for the same country/location/pollutant/year.",
        )

    upload = create_upload(
        db,
        account_id=account.account_id,
        org_id=org.org_id,
        data_domain=org.data_domain,
        country=org.country,
        data=UploadCreate(mongo_collection="OpenAQ", mongo_ref_id=batch_id),
    )
    update_upload_status(db, upload, UploadUpdateStatus(status=UploadStatus.PROCESSED))
    return upload


def list_upload_records(db: Session, account: Account, upload_id: int, limit: int, offset: int):
    upload = get_upload_by_id(db, upload_id)
    if not upload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload not found")
    if account.role == AccountRole.ORG and account.org_id != upload.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")

    batch_id = ObjectId(upload.mongo_ref_id)
    if upload.mongo_collection == "OpenAQ":
        col = get_openaq_collection()
        query = {"_source_batch": batch_id}
    else:
        col = get_imhe_collection()
        query = {"_source_batch": batch_id}
    total = col.count_documents(query)
    cursor = col.find(query).skip(int(offset)).limit(int(limit))
    items = []
    for doc in cursor:
        doc["id"] = str(doc.pop("_id"))
        doc.pop("_source_batch", None)
        items.append(doc)
    return {"total": total, "items": items}


def update_upload_record(
    db: Session,
    account: Account,
    upload_id: int,
    record_id: str,
    payload: UploadRecordUpdate,
):
    upload = get_upload_by_id(db, upload_id)
    if not upload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload not found")
    if account.role == AccountRole.ORG and account.org_id != upload.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")
    if upload.mongo_collection == "OpenAQ":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Pollution records are read-only")
    # Validate year and required fields
    current_year = datetime.utcnow().year
    if payload.year < 1900 or payload.year > current_year + 3:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"year must be between 1900 and {current_year + 3}",
        )
    if payload.val is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="val is required")
    if payload.upper is not None and payload.lower is not None:
        if payload.lower > payload.upper:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="lower cannot exceed upper")

    col = get_imhe_collection()
    batch_id = ObjectId(upload.mongo_ref_id)
    doc_id = ObjectId(record_id)

    update_doc = {
        "measure_name": payload.measure_name,
        "sex_name": payload.sex_name,
        "age_name": payload.age_name,
        "cause_name": payload.cause_name,
        "metric_name": payload.metric_name,
        "year": payload.year,
        "val": payload.val,
        "upper": payload.upper if payload.upper is not None else payload.val,
        "lower": payload.lower if payload.lower is not None else payload.val,
    }
    update_doc["measure_id"] = _resolve_id(col, "measure_id", "measure_name", payload.measure_name)
    update_doc["sex_id"] = _resolve_id(col, "sex_id", "sex_name", payload.sex_name)
    update_doc["age_id"] = _resolve_id(col, "age_id", "age_name", payload.age_name)
    update_doc["cause_id"] = _resolve_id(col, "cause_id", "cause_name", payload.cause_name)
    update_doc["metric_id"] = _resolve_id(col, "metric_id", "metric_name", payload.metric_name)

    result = col.update_one({"_id": doc_id, "_source_batch": batch_id}, {"$set": update_doc})
    if result.matched_count == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Record not found")
    return {"status": "ok"}


def delete_upload_with_records(db: Session, account: Account, upload_id: int):
    upload = get_upload_by_id(db, upload_id)
    if not upload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload not found")
    if account.role == AccountRole.ORG and account.org_id != upload.org_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")

    col = get_openaq_collection() if upload.mongo_collection == "OpenAQ" else get_imhe_collection()
    try:
        batch_id = ObjectId(upload.mongo_ref_id)
        col.delete_many({"_source_batch": batch_id})
    except Exception:
        pass
    delete_upload(db, upload)
    return {"status": "deleted"}


def list_uploads_for_account(db: Session, account: Account):
    if account.role == AccountRole.ADMIN:
        return list_uploads(db)
    if account.role == AccountRole.ORG and account.org_id:
        return list_uploads_by_org(db, account.org_id)
    return []


def admin_update_upload(db: Session, upload_id: int, data: UploadUpdateStatus):
    upload = get_upload_by_id(db, upload_id)
    if not upload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Upload not found")
    return update_upload_status(db, upload, data)
